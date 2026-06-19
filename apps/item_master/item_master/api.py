"""
item_master/item_master/api.py
Server-side API cho chức năng upload master data sản phẩm.
"""

import frappe
from frappe import _
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Cấu trúc template ────────────────────────────────────────────────────────
TEMPLATE_COLUMNS = [
    # (field_name, header_vi, required, example, width)
    ("item_code",          "Mã sản phẩm *",         True,  "SP001",           18),
    ("item_name",          "Tên sản phẩm *",         True,  "Sản phẩm mẫu",   28),
    ("item_group",         "Nhóm sản phẩm *",        True,  "Thành phẩm",     22),
    ("description",        "Mô tả",                  False, "",                30),
    ("stock_uom",          "Đơn vị tính *",          True,  "Cái",             14),
    ("is_stock_item",      "Quản lý tồn kho (1/0)",  False, "1",               22),
    ("has_variants",       "Có biến thể (1/0)",      False, "0",               20),
    ("standard_rate",      "Giá bán",                False, "100000",          14),
    ("valuation_rate",     "Giá vốn",                False, "80000",           14),
    ("weight_per_unit",    "Trọng lượng",             False, "1.5",             14),
    ("weight_uom",         "ĐVT trọng lượng",         False, "Kg",              16),
    ("brand",              "Thương hiệu",             False, "",                16),
    ("disabled",           "Vô hiệu hóa (1/0)",      False, "0",               20),
]


@frappe.whitelist()
def download_template():
    """Tạo và trả về file Excel template cho upload Item master data."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Template nhập liệu ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Dữ liệu sản phẩm"

    # Style định nghĩa
    header_fill   = PatternFill("solid", fgColor="1F4E79")
    required_fill = PatternFill("solid", fgColor="2E75B6")
    header_font   = Font(color="FFFFFF", bold=True, size=11)
    required_font = Font(color="FFFF00", bold=True, size=11)
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align    = Alignment(horizontal="left",   vertical="center")
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    # Dòng tiêu đề
    for col_idx, (_, header, required, _, width) in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill  = required_fill if required else header_fill
        cell.font  = required_font if required else header_font
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 35

    # Dòng ví dụ
    example_row = [ex for _, _, _, ex, _ in TEMPLATE_COLUMNS]
    for col_idx, value in enumerate(example_row, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.fill      = PatternFill("solid", fgColor="EBF3FB")
        cell.alignment = left_align
        cell.border    = thin_border

    # Đóng băng dòng tiêu đề
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Hướng dẫn ──────────────────────────────────────────────────
    ws_guide = wb.create_sheet("Hướng dẫn")
    guide_content = [
        ("Hướng dẫn sử dụng template",),
        ("",),
        ("1. Cột màu xanh đậm là BẮT BUỘC (Mã SP, Tên SP, Nhóm SP, ĐVT)",),
        ("2. Mã sản phẩm phải là duy nhất trong hệ thống",),
        ("3. Nhóm sản phẩm phải tồn tại trong danh mục (vào ERPNext > Item Group)",),
        ("4. Đơn vị tính phải tồn tại trong UOM (vào ERPNext > UOM)",),
        ("5. Cột số nhị phân (1/0): 1 = Có, 0 = Không",),
        ("6. Giá nhập theo đơn vị tiền tệ mặc định của hệ thống",),
        ("7. Dòng đầu tiên là ví dụ minh họa, có thể xóa hoặc giữ lại",),
        ("8. Không thay đổi tiêu đề cột",),
        ("",),
        ("Nhóm sản phẩm thông dụng:",),
        ("  - Thành phẩm",),
        ("  - Nguyên vật liệu",),
        ("  - Hàng hóa",),
        ("  - Dịch vụ",),
        ("",),
        ("Đơn vị tính thông dụng: Cái, Kg, Lit, Thùng, Bộ, Hộp, Mét, m2, m3",),
    ]
    for row_idx, row_data in enumerate(guide_content, 1):
        cell = ws_guide.cell(row=row_idx, column=1, value=row_data[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=13, color="1F4E79")
        elif row_data[0].startswith(("1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.")):
            cell.font = Font(size=11)
        else:
            cell.font = Font(size=10, color="595959")
    ws_guide.column_dimensions["A"].width = 70

    # ── Xuất file ────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    frappe.response.filename    = "Template_SanPham.xlsx"
    frappe.response.filecontent = buffer.read()
    frappe.response.type        = "binary"


def _ensure_uoms_exist(data_rows, field_indices):
    """Tạo bản ghi UOM cho các đơn vị chưa có trong hệ thống."""
    uom_fields = ["stock_uom", "weight_uom"]
    needed = set()
    for row in data_rows:
        if not any(row):
            continue
        for field in uom_fields:
            idx = field_indices.get(field)
            if idx is not None and idx < len(row) and row[idx]:
                val = str(row[idx]).strip()
                if val:
                    needed.add(val)

    for uom in needed:
        if not frappe.db.exists("UOM", uom):
            doc = frappe.new_doc("UOM")
            doc.uom_name = uom
            doc.insert(ignore_permissions=True)

    if needed:
        frappe.db.commit()


@frappe.whitelist()
def upload_master_data(file_url, update_existing=1, skip_errors=0):
    """
    Nhập Items từ file Excel đã upload.
    Args:
        file_url:        URL của file đã upload (từ Frappe file manager)
        update_existing: Cập nhật item nếu đã tồn tại
        skip_errors:     Bỏ qua dòng lỗi, tiếp tục xử lý
    Returns:
        dict: { created, updated, skipped, errors }
    """
    if not set(frappe.get_roles(frappe.session.user)) & {"System Manager", "Item Manager", "Stock Manager"}:
        frappe.throw(frappe._("Bạn không có quyền thực hiện thao tác này"), frappe.PermissionError)

    update_existing = int(update_existing)
    skip_errors     = int(skip_errors)

    # Đọc file từ Frappe file manager
    _file = frappe.get_doc("File", {"file_url": file_url})
    file_path = _file.get_full_path()

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        frappe.throw(_("File trống"))

    # Đọc header từ dòng 1
    headers   = [str(h).strip() if h else "" for h in rows[0]]
    col_map   = {col[0]: idx for idx, col in enumerate(TEMPLATE_COLUMNS)
                 if col[0] in [c[0] for c in TEMPLATE_COLUMNS]}

    # Map tên header → field name
    header_to_field = {col[1]: col[0] for col in TEMPLATE_COLUMNS}
    field_indices   = {}
    for h_idx, header in enumerate(headers):
        clean_header = header.replace(" *", "").strip()
        for col_header, field in header_to_field.items():
            if col_header.replace(" *", "").strip() == clean_header:
                field_indices[field] = h_idx
                break

    # Tạo UOM còn thiếu trước khi xử lý từng dòng
    _ensure_uoms_exist(rows[2:], field_indices)

    results = {"created": 0, "updated": 0, "skipped": 0, "errors": []}

    for row_num, row in enumerate(rows[2:], start=3):  # Bỏ header + ví dụ
        if not any(row):
            continue

        def get(field):
            idx = field_indices.get(field)
            if idx is None:
                return None
            val = row[idx] if idx < len(row) else None
            return str(val).strip() if val is not None else None

        item_code = get("item_code")
        item_name = get("item_name")
        item_group = get("item_group")
        stock_uom  = get("stock_uom")

        # Validate bắt buộc
        missing = []
        if not item_code:  missing.append("Mã sản phẩm")
        if not item_name:  missing.append("Tên sản phẩm")
        if not item_group: missing.append("Nhóm sản phẩm")
        if not stock_uom:  missing.append("Đơn vị tính")

        if missing:
            msg = f"Dòng {row_num}: Thiếu {', '.join(missing)}"
            if skip_errors:
                results["errors"].append(msg)
                results["skipped"] += 1
                continue
            else:
                frappe.throw(msg)

        exists = frappe.db.exists("Item", item_code)

        if exists and not update_existing:
            results["skipped"] += 1
            continue

        try:
            if exists:
                doc = frappe.get_doc("Item", item_code)
            else:
                doc = frappe.new_doc("Item")
                doc.item_code = item_code

            doc.item_name  = item_name
            doc.item_group = item_group
            doc.stock_uom  = stock_uom

            # Fields tuỳ chọn
            if get("description"):
                doc.description = get("description")
            if get("is_stock_item") is not None:
                doc.is_stock_item = int(get("is_stock_item") or 1)
            if get("has_variants") is not None:
                doc.has_variants = int(get("has_variants") or 0)
            if get("standard_rate"):
                doc.standard_rate = float(get("standard_rate"))
            if get("valuation_rate"):
                doc.valuation_rate = float(get("valuation_rate"))
            if get("weight_per_unit"):
                doc.weight_per_unit = float(get("weight_per_unit"))
            if get("weight_uom"):
                doc.weight_uom = get("weight_uom")
            if get("brand"):
                doc.brand = get("brand")
            if get("disabled") is not None:
                doc.disabled = int(get("disabled") or 0)

            doc.save(ignore_permissions=False)
            frappe.db.commit()

            if exists:
                results["updated"] += 1
            else:
                results["created"] += 1

        except Exception as e:
            msg = f"Dòng {row_num} ({item_code}): {str(e)}"
            if skip_errors:
                results["errors"].append(msg)
                results["skipped"] += 1
                frappe.db.rollback()
            else:
                frappe.throw(msg)

    wb.close()
    return results
